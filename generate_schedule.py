import csv
import random
from datetime import datetime, timedelta, time
from collections import defaultdict
from pathlib import Path

try:
    import pandas as pd  # optional for Excel output
except Exception:  # pragma: no cover
    pd = None


# ============================ Konfigurasi ============================
START_DATE = datetime(2025, 11, 3)  # mulai 3 Nov 2025
END_DATE = datetime(2025, 11, 7)    # berakhir 7 Nov 2025 (weekday only)
DAY_START = time(7, 30)
DAY_END = time(17, 30)
SHIFT_DURATION_MIN = 120  # 2 jam
GAP_BETWEEN_SHIFTS_MIN = 30  # jeda antar shift
LUNCH_BREAK_START = time(12, 0)
LUNCH_BREAK_END = time(13, 0)

# Shifts resmi (tetap): 07.30-09.30, 10.00-12.00, 13.00-15.00, 15.30-17.30
ALLOWED_SHIFT_STARTS = [time(7, 30), time(10, 0), time(13, 0), time(15, 30)]

# Daftar ruangan yang tersedia - akan dimuat dari ruangan-kampus.csv
ALL_ROOMS = []

# Blacklist ruangan berdasarkan hari (khusus minggu UTS 3-7 Nov 2025)
BLACKLIST_MON_WED_SUFFIXES = {"KELAS 2.08", "KELAS 2.07", "KELAS 2.06", "KELAS 2.05", "KELAS 2.04"}
BLACKLIST_MON_FRI_SUFFIXES = {"KELAS 2.09"}


def is_within_uts_week(date_dt: datetime) -> bool:
    return START_DATE.date() <= date_dt.date() <= END_DATE.date()


def is_room_blacklisted_on_date(room: str, date_dt: datetime) -> bool:
    """Return True jika ruangan diblacklist pada tanggal tersebut."""
    if not is_within_uts_week(date_dt):
        return False
    weekday = date_dt.weekday()  # 0=Mon ... 4=Fri
    # KELAS 2.09 diblacklist Senin-Jumat
    for suf in BLACKLIST_MON_FRI_SUFFIXES:
        if room.endswith(suf) and weekday <= 4:
            return True
    # Lainnya diblacklist Senin-Rabu
    for suf in BLACKLIST_MON_WED_SUFFIXES:
        if room.endswith(suf) and weekday <= 2:
            return True
    return False


# ============================ Aturan Khusus AULA ============================
AULA_NAME = "AULA"

def is_aula(room: str) -> bool:
    return room.strip().upper() == AULA_NAME

def is_aula_time_allowed(date_dt: datetime, start_dt: datetime, end_dt: datetime) -> bool:
    """AULA rules:
    - HANYA Senin dan Selasa di minggu UTS.
      - Senin (2025-11-03): boleh full 07.30 - 17.30
      - Selasa (2025-11-04): hanya boleh mulai dari 13.00 ke atas (start >= 13.00)
    - Hari lain: TIDAK BOLEH dipakai
    """
    if not is_within_uts_week(date_dt):
        return False
    weekday = date_dt.weekday()  # 0=Mon, 1=Tue
    if weekday == 0:  # Senin
        return True
    if weekday == 1:  # Selasa
        allowed_start = datetime.combine(date_dt.date(), time(13, 0))
        return start_dt >= allowed_start
    # Rabu-Kamis-Jumat: tidak boleh
    return False

def aula_preferred_shifts(day_dt: datetime) -> list[tuple[datetime, datetime]]:
    """Return shifts for a date reordered to prefer afternoon first for AULA.
    For Tuesday: 13:00, 15:30, then 07:30, 10:00.
    For Monday: 07:30, 10:00, then 13:00, 15:30.
    """
    slots = generate_daily_shifts(day_dt)
    # map for easy lookup
    if day_dt.weekday() == 0:  # Monday
        order = [time(7, 30), time(10, 0), time(13, 0), time(15, 30)]
    else:
        # Tuesday and others
        order = [time(13, 0), time(15, 30), time(7, 30), time(10, 0)]
    by_start = {s[0].time(): s for s in slots}
    return [by_start[t] for t in order if t in by_start]

def aula_preferred_dates() -> list[datetime]:
    """Prefer Tuesday first, then Monday within the UTS week."""
    # Build list of allowed dates then reorder
    dates = list(iter_allowed_dates())
    mon = [d for d in dates if d.weekday() == 0]  # Monday
    tue = [d for d in dates if d.weekday() == 1]  # Tuesday
    others = [d for d in dates if d.weekday() not in (0, 1)]
    return mon + tue + others


def parse_csv(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=";")
        rows = list(reader)
    # Header ada pada baris pertama file
    header = rows[0]
    # Buat mapping kolom ke index
    col_idx = {name.strip().upper(): i for i, name in enumerate(header)}

    def get(row, key, default=""):
        i = col_idx.get(key, None)
        if i is None or i >= len(row):
            return default
        return row[i].strip()

    items = []
    for r in rows[1:]:
        if not any(r):
            continue
        kode_mk = get(r, "KODE MATA KULIAH")
        nama_mk = get(r, "NAMA MATA KULIAH")
        nama_dosen = get(r, "NAMA DOSEN")
        kelas = get(r, "KELAS")
        if not kode_mk and not nama_mk:
            continue
        hari = get(r, "HARI")
        tanggal = get(r, "TANGGAL")
        shift = get(r, "SHIFT")  # contoh: "07.30 - 09.30"
        ruangan = get(r, "RUANGAN")
        bentuk_ujian = get(r, "BENTUK UJIAN")
        butuh_gandakan = get(r, "BUTUH MENGGANDAKAN SOAL")
        butuh_lembar = get(r, "BUTUH LEMBAR JAWABAN KERJA")
        butuh_pengawas = get(r, "BUTUH PENGAWAS UJIAN")
        butuh_ruang = get(r, "BUTUH RUANG KELAS")
        jumlah_mhs = get(r, "JUMLAH MAHASISWA")

        items.append({
            "kode_mk": kode_mk,
            "nama_mk": nama_mk,
            "nama_dosen": nama_dosen,
            "kelas": kelas,
            "hari": hari,
            "tanggal": tanggal,
            "shift": shift,
            "ruangan": ruangan,
            "bentuk_ujian": bentuk_ujian,
            "butuh_gandakan": butuh_gandakan.upper() if butuh_gandakan else "",
            "butuh_lembar": butuh_lembar.upper() if butuh_lembar else "",
            "butuh_pengawas": butuh_pengawas.upper() if butuh_pengawas else "",
            "butuh_ruang": butuh_ruang.upper() if butuh_ruang else "",
            "jumlah_mhs": jumlah_mhs,
        })
    return items


def load_rooms_from_csv(rooms_csv_path: Path) -> list[str]:
    """Load room names from ruangan-kampus.csv file."""
    rooms = []
    try:
        with rooms_csv_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, delimiter=";")
            rows = list(reader)
        
        # Skip header row (first row)
        for row in rows[1:]:
            if len(row) > 0 and row[0].strip():  # Check if RUANGAN column has data
                room_name = row[0].strip()
                if room_name:  # Only add non-empty room names
                    rooms.append(room_name)
    except Exception as e:
        print(f"Error loading rooms from {rooms_csv_path}: {e}")
        print("Using empty room list as fallback")
    
    return rooms


def format_time_range(start_dt: datetime, end_dt: datetime) -> str:
    return f"{start_dt:%H.%M} - {end_dt:%H.%M}"


def generate_daily_shifts(start_date: datetime) -> list[tuple[datetime, datetime]]:
    """Kembalikan daftar shift tetap untuk tanggal tersebut."""
    shifts: list[tuple[datetime, datetime]] = []
    for s in ALLOWED_SHIFT_STARTS:
        start_dt = datetime.combine(start_date.date(), s)
        end_dt = start_dt + timedelta(minutes=SHIFT_DURATION_MIN)
        shifts.append((start_dt, end_dt))
    return shifts


def normalize_to_allowed_shift(date_dt: datetime, start_dt: datetime) -> tuple[datetime, datetime]:
    """Map arbitrary start time to the nearest allowed 2-hour shift on that date.
    Allowed starts: 07:30, 10:00, 13:00, 15:30
    """
    allowed = [s for s in generate_daily_shifts(date_dt)]
    if not allowed:
        return start_dt, start_dt + timedelta(minutes=SHIFT_DURATION_MIN)
    # Choose allowed slot with minimal absolute difference in start time
    best = min(allowed, key=lambda se: abs((se[0] - start_dt).total_seconds()))
    return best


def weekday_name(dt: datetime) -> str:
    mapping = {
        0: "SENIN",
        1: "SELASA",
        2: "RABU",
        3: "KAMIS",
        4: "JUM'AT",
        5: "SABTU",
        6: "MINGGU",
    }
    return mapping[dt.weekday()]


def iter_allowed_dates():
    """Iterasi tanggal dari START_DATE s.d END_DATE, hanya Senin-Jumat."""
    cur = START_DATE
    while cur.date() <= END_DATE.date():
        if cur.weekday() < 5:  # 0-4: Mon-Fri
            yield cur
        cur += timedelta(days=1)


def parse_existing_datetime(hari: str, tanggal: str, shift: str) -> tuple[datetime, datetime] | None:
    if not tanggal or not shift:
        return None
    # Coba beberapa format tanggal
    date_formats = ["%d-%b-%y", "%d/%m/%Y", "%d-%m-%Y"]
    date_dt = None
    for fmt in date_formats:
        try:
            date_dt = datetime.strptime(tanggal.strip(), fmt)
            break
        except Exception:
            continue
    if date_dt is None:
        return None
    # shift contoh: 07.30 - 09.30 (bisa ada tab/extra space)
    parts = shift.split("-")
    if len(parts) != 2:
        return None
    start_s = parts[0].strip().replace(" ", "").replace("\t", "")
    end_s = parts[1].strip().replace(" ", "").replace("\t", "")
    try:
        h1, m1 = map(int, start_s.split("."))
        h2, m2 = map(int, end_s.split("."))
    except Exception:
        return None
    start_dt = datetime.combine(date_dt.date(), time(h1, m1))
    end_dt = datetime.combine(date_dt.date(), time(h2, m2))
    return start_dt, end_dt


def build_schedule(items: list[dict]):
    # State pemakaian: per (tanggal, shift_str) -> room->count pemakaian, dan kelas-> list times
    room_usage = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))  # date_str -> shift_str -> room -> count
    room_occupants = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))  # date_str -> shift_str -> room -> list kelas
    class_usage = defaultdict(list)  # kelas -> list[(start_dt,end_dt)]
    class_daily_count = defaultdict(lambda: defaultdict(int))  # kelas -> date_str -> count

    # Kumpulkan existing jadwal bila ada
    for it in items:
        if it["tanggal"] and it["shift"]:
            parsed = parse_existing_datetime(it["hari"], it["tanggal"], it["shift"]) 
            if parsed is None:
                continue
            start_dt, end_dt = parsed
            date_key = start_dt.strftime("%Y-%m-%d")
            shift_key = format_time_range(start_dt, end_dt)
            cls = it["kelas"]
            if cls:
                class_usage[cls].append((start_dt, end_dt))
                class_daily_count[cls][date_key] += 1
            room = it["ruangan"].strip() if it["ruangan"] else ""
            if room:
                room_usage[date_key][shift_key][room] += 1
                cls_for_occ = it.get("kelas", "")
                if cls_for_occ:
                    room_occupants[date_key][shift_key][room].append(cls_for_occ)

    # Siapkan daftar shifts harian (dipakai untuk mengisi yang kosong)
    # Kita generate untuk beberapa hari ke depan (misal 14 hari) sampai kebutuhan terpenuhi
    generated_assignments: list[dict] = []

    def is_class_conflict(kelas: str, start_dt: datetime, end_dt: datetime) -> bool:
        if not kelas:
            return False
        
        # Check for time overlap conflicts
        for s, e in class_usage.get(kelas, []):
            if not (end_dt <= s or start_dt >= e):
                return True
        
        # Check for daily limit (max 2 exams per day)
        date_key = start_dt.strftime("%Y-%m-%d")
        if class_daily_count[kelas][date_key] >= 2:
            return True
            
        return False

    # Fungsi memilih ruangan kosong pada tanggal+shift tertentu (memperhatikan blacklist per tanggal)
    def pick_free_room(date_dt: datetime, date_key: str, shift_key: str, start_dt: datetime, end_dt: datetime, bentuk_ujian: str, allow_aula: bool, jumlah_mhs: int = 0) -> str | None:
        used_counts = room_usage[date_key][shift_key]
        aula_candidates = []
        normal_candidates = []
        for r in ALL_ROOMS:
            if is_room_blacklisted_on_date(r, date_dt):
                continue
            count = used_counts.get(r, 0)
            if is_aula(r):
                if not allow_aula:
                    continue
                # AULA hanya untuk ujian tulis dengan jumlah_mhs diketahui (>0)
                bentuk = (bentuk_ujian or "").strip().lower()
                if bentuk != "ujian tulis" or jumlah_mhs <= 0 or jumlah_mhs < 40:
                    continue
                # AULA boleh hingga 2 kelas per shift, dan patuhi aturan waktu khusus
                if count < 2 and is_aula_time_allowed(date_dt, start_dt, end_dt):
                    aula_candidates.append(r)
            else:
                if count == 0:
                    normal_candidates.append(r)
        if not aula_candidates and not normal_candidates:
            return None
        bentuk = (bentuk_ujian or "").strip().lower()
        is_tulis = bentuk == "ujian tulis" and jumlah_mhs > 0 and jumlah_mhs >= 40
        if is_tulis:
            # Prioritaskan AULA dulu
            if aula_candidates:
                return AULA_NAME if AULA_NAME in aula_candidates else random.choice(aula_candidates)
            if normal_candidates:
                return random.choice(normal_candidates)
        else:
            # Non-"Ujian Tulis": gunakan ruangan biasa dulu, AULA sebagai fallback
            if normal_candidates:
                return random.choice(normal_candidates)
            if aula_candidates:
                return AULA_NAME if AULA_NAME in aula_candidates else random.choice(aula_candidates)
        return None

    # Helper: iterate allowed dates and shifts until assignable

    # Sort items by jumlah mahasiswa desc to prioritize bigger classes for AULA
    def parse_int_safe(v: str) -> int:
        try:
            return int(str(v).strip())
        except Exception:
            return 0

    items_with_index = list(enumerate(items))

    def sort_key(entry):
        _, it0 = entry
        bentuk0 = (it0.get("bentuk_ujian", "") or "").strip().lower()
        is_tulis0 = bentuk0 == "ujian tulis"
        has_shift0 = bool(
            (it0.get("hari") or "").strip()
            and (it0.get("tanggal") or "").strip()
            and (it0.get("shift") or "").strip()
        )
        kelas0 = (it0.get("kelas", "") or "").strip()
        prefix0 = kelas0[:2]
        mhs0 = parse_int_safe(it0.get("jumlah_mhs", "0"))
        # AULA candidate: Ujian Tulis and no pre-defined shift
        aula_cand = (is_tulis0 and not has_shift0)
        # Sort: AULA candidates first (1), then prefix, then jumlah desc
        return (1 if aula_cand else 0, prefix0, -mhs0)

    items_with_index.sort(key=sort_key, reverse=False)

    for _, it in items_with_index:
        kode = it["kode_mk"]
        nama = it["nama_mk"]
        kelas = it["kelas"]
        hari = it["hari"].strip().upper() if it["hari"] else ""
        tanggal = it["tanggal"].strip() if it["tanggal"] else ""
        shift = it["shift"].strip() if it["shift"] else ""
        ruangan = it["ruangan"].strip() if it["ruangan"] else ""
        bentuk_ujian = (it.get("bentuk_ujian", "") or "").strip()
        jumlah_mhs_val = parse_int_safe(it.get("jumlah_mhs", "0"))

        # Jika semua field (hari, tanggal, shift, ruangan) sudah terisi di CSV,
        # gunakan persis apa adanya (TIDAK dinormalisasi), tapi tetap catat ke state bila bisa di-parse.
        if hari and tanggal and shift and ruangan:
            parsed_full = parse_existing_datetime(hari, tanggal, shift)
            if parsed_full is not None:
                s_dt, e_dt = parsed_full
                date_key = s_dt.strftime("%Y-%m-%d")
                shift_key_state = format_time_range(s_dt, e_dt)
                if kelas:
                    class_usage[kelas].append((s_dt, e_dt))
                    class_daily_count[kelas][date_key] += 1
                room = ruangan.strip()
                if room:
                    room_usage[date_key][shift_key_state][room] += 1
                    if kelas:
                        room_occupants[date_key][shift_key_state][room].append(kelas)
            # Tulis PERSIS seperti CSV
            generated_assignments.append({
                "HARI": it["hari"],
                "TANGGAL": it["tanggal"],
                "SHIFT": it["shift"],
                "RUANGAN": it["ruangan"],
                "KODE MATA KULIAH": kode,
                "NAMA MATA KULIAH": nama,
                "NAMA DOSEN": it.get("nama_dosen", ""),
                "KELAS": kelas,
                "BENTUK UJIAN": it.get("bentuk_ujian", ""),
                "BUTUH MENGGANDAKAN SOAL": it.get("butuh_gandakan", ""),
                "BUTUH LEMBAR JAWABAN KERJA": it.get("butuh_lembar", ""),
                "BUTUH PENGAWAS UJIAN": it.get("butuh_pengawas", ""),
                "BUTUH RUANG KELAS": it.get("butuh_ruang", ""),
                "JUMLAH MAHASISWA": it.get("jumlah_mhs", ""),
            })
            continue

        # Jika hari, tanggal, shift sudah ada: JANGAN ubah waktu. Hanya carikan ruangan jika kosong.
        parsed = parse_existing_datetime(hari, tanggal, shift)
        if parsed is not None:
            start_dt, end_dt = parsed
            date_key = start_dt.strftime("%Y-%m-%d")
            shift_key = format_time_range(start_dt, end_dt)
            # Jika ruangan sudah ada di CSV, pakai apa adanya (tidak dirandom)
            # Jika kosong, baru cari ruangan kosong secara acak
            room = ruangan if ruangan else pick_free_room(start_dt, date_key, shift_key, start_dt, end_dt, bentuk_ujian, False, jumlah_mhs_val)
            if room is None:
                room = ""
            if kelas:
                class_usage[kelas].append((start_dt, end_dt))
                class_daily_count[kelas][date_key] += 1
            if room:
                room_usage[date_key][shift_key][room] += 1
                if kelas:
                    room_occupants[date_key][shift_key][room].append(kelas)
            generated_assignments.append({
                "HARI": weekday_name(start_dt),
                "TANGGAL": start_dt.strftime("%d-%b-%y"),
                "SHIFT": shift_key,
                "RUANGAN": room,
                "KODE MATA KULIAH": kode,
                "NAMA MATA KULIAH": nama,
                "NAMA DOSEN": it.get("nama_dosen", ""),
                "KELAS": kelas,
                "BENTUK UJIAN": it.get("bentuk_ujian", ""),
                "BUTUH MENGGANDAKAN SOAL": it.get("butuh_gandakan", ""),
                "BUTUH LEMBAR JAWABAN KERJA": it.get("butuh_lembar", ""),
                "BUTUH PENGAWAS UJIAN": it.get("butuh_pengawas", ""),
                "BUTUH RUANG KELAS": it.get("butuh_ruang", ""),
                "JUMLAH MAHASISWA": it.get("jumlah_mhs", ""),
            })
            continue

        # Jika belum ada hari/tanggal/shift, generate baru
        assigned = False

        # 1) Coba pairing ke AULA yang sudah punya 1 slot terisi terlebih dahulu (prefer prefix sama)
        #    Hanya untuk BENTUK UJIAN = "Ujian Tulis"
        #    Hanya untuk tanggal/shift yang kompatibel dengan kelas ini (tidak konflik) dan aturan AULA.
        for day_dt in aula_preferred_dates():
            for s_start, s_end in aula_preferred_shifts(day_dt):
                if is_class_conflict(kelas, s_start, s_end):
                    continue
                date_key = s_start.strftime("%Y-%m-%d")
                shift_key = format_time_range(s_start, s_end)
                # Cari AULA yang masih count < 2 dan waktu diizinkan
                counts = room_usage[date_key][shift_key]
                aula_count = counts.get(AULA_NAME, 0)
                if (bentuk_ujian.strip().lower() == "ujian tulis") and jumlah_mhs_val >= 40 and aula_count == 1 and is_aula_time_allowed(day_dt, s_start, s_end):
                    # Prefer jika prefix kelas sama
                    occupants = room_occupants[date_key][shift_key][AULA_NAME]
                    prefer = False
                    if occupants:
                        try:
                            pref_new = (kelas or "")[:2]
                            pref_old = (occupants[0] or "")[:2]
                            prefer = bool(pref_new) and bool(pref_old) and pref_new == pref_old
                        except Exception:
                            prefer = False
                    # Wajib sama prefix untuk mengisi slot kedua AULA
                    if not prefer:
                        continue
                    
                    room = AULA_NAME
                    class_usage[kelas].append((s_start, s_end))
                    class_daily_count[kelas][date_key] += 1
                    room_usage[date_key][shift_key][room] += 1
                    if kelas:
                        room_occupants[date_key][shift_key][room].append(kelas)
                    generated_assignments.append({
                        "HARI": weekday_name(s_start),
                        "TANGGAL": s_start.strftime("%d-%b-%y"),
                        "SHIFT": shift_key,
                        "RUANGAN": room,
                        "KODE MATA KULIAH": kode,
                        "NAMA MATA KULIAH": nama,
                        "NAMA DOSEN": it.get("nama_dosen", ""),
                        "KELAS": kelas,
                        "BENTUK UJIAN": it.get("bentuk_ujian", ""),
                        "BUTUH MENGGANDAKAN SOAL": it.get("butuh_gandakan", ""),
                        "BUTUH LEMBAR JAWABAN KERJA": it.get("butuh_lembar", ""),
                        "BUTUH PENGAWAS UJIAN": it.get("butuh_pengawas", ""),
                        "BUTUH RUANG KELAS": it.get("butuh_ruang", ""),
                        "JUMLAH MAHASISWA": it.get("jumlah_mhs", ""),
                    })
                    assigned = True
                    break
            if assigned:
                break

        if assigned:
            continue

        # 2) Alokasi normal (memungkinkan AULA dengan kapasitas 2)
        is_aula_candidate = (bentuk_ujian.strip().lower() == "ujian tulis" and jumlah_mhs_val > 0)
        date_iter = aula_preferred_dates() if is_aula_candidate else iter_allowed_dates()
        for day_dt in date_iter:
            shift_iter = aula_preferred_shifts(day_dt) if is_aula_candidate else generate_daily_shifts(day_dt)
            for s_start, s_end in shift_iter:
                if is_class_conflict(kelas, s_start, s_end):
                    continue
                date_key = s_start.strftime("%Y-%m-%d")
                shift_key = format_time_range(s_start, s_end)
                # Jika CSV sudah menspesifikkan ruangan, coba pakai ruangan itu saja
                if ruangan:
                    # Hanya assign jika ruangan tersebut belum dipakai pada slot ini
                    if not is_room_blacklisted_on_date(ruangan, s_start):
                        current = room_usage[date_key][shift_key].get(ruangan, 0)
                        if is_aula(ruangan):
                            if is_aula_time_allowed(day_dt, s_start, s_end) and current < 2:
                                room = ruangan
                            else:
                                continue
                        else:
                            if current == 0:
                                room = ruangan
                            else:
                                continue
                    else:
                        # Slot ini tidak tersedia untuk ruangan yang diminta, coba slot berikutnya
                        continue
                else:
                    room = pick_free_room(s_start, date_key, shift_key, s_start, s_end, bentuk_ujian, True, jumlah_mhs_val)
                if room:
                    class_usage[kelas].append((s_start, s_end))
                    class_daily_count[kelas][date_key] += 1
                    room_usage[date_key][shift_key][room] += 1
                    if kelas:
                        room_occupants[date_key][shift_key][room].append(kelas)
                    generated_assignments.append({
                        "HARI": weekday_name(s_start),
                        "TANGGAL": s_start.strftime("%d-%b-%y"),
                        "SHIFT": shift_key,
                        "RUANGAN": room,
                        "KODE MATA KULIAH": kode,
                        "NAMA MATA KULIAH": nama,
                        "NAMA DOSEN": it.get("nama_dosen", ""),
                        "KELAS": kelas,
                        "BENTUK UJIAN": it.get("bentuk_ujian", ""),
                        "BUTUH MENGGANDAKAN SOAL": it.get("butuh_gandakan", ""),
                        "BUTUH LEMBAR JAWABAN KERJA": it.get("butuh_lembar", ""),
                        "BUTUH PENGAWAS UJIAN": it.get("butuh_pengawas", ""),
                        "BUTUH RUANG KELAS": it.get("butuh_ruang", ""),
                        "JUMLAH MAHASISWA": it.get("jumlah_mhs", ""),
                    })
                    assigned = True
                    break
            if assigned:
                break
        if not assigned:
            # gagal assign, tetap keluarkan tanpa ruangan
            generated_assignments.append({
                "HARI": "",
                "TANGGAL": "",
                "SHIFT": "",
                "RUANGAN": "",
                "KODE MATA KULIAH": kode,
                "NAMA MATA KULIAH": nama,
                "NAMA DOSEN": it.get("nama_dosen", ""),
                "KELAS": kelas,
                "BENTUK UJIAN": it.get("bentuk_ujian", ""),
                "BUTUH MENGGANDAKAN SOAL": it.get("butuh_gandakan", ""),
                "BUTUH LEMBAR JAWABAN KERJA": it.get("butuh_lembar", ""),
                "BUTUH PENGAWAS UJIAN": it.get("butuh_pengawas", ""),
                "BUTUH RUANG KELAS": it.get("butuh_ruang", ""),
                "JUMLAH MAHASISWA": it.get("jumlah_mhs", ""),
            })

    return generated_assignments


def write_outputs(assignments, out_csv: Path, out_xlsx: Path | None):
    # Kolom output sesuai permintaan
    cols = [
        "HARI",
        "TANGGAL",
        "SHIFT",
        "RUANGAN",
        "KODE MATA KULIAH",
        "NAMA MATA KULIAH",
        "NAMA DOSEN",
        "KELAS",
        "BENTUK UJIAN",
        "JUMLAH MAHASISWA",
    ]
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for row in assignments:
            w.writerow([row.get(c, "") for c in cols])

    if pd is not None and out_xlsx is not None:
        try:
            # Tuliskan menggunakan pandas
            df = pd.DataFrame(assignments, columns=cols)
            # Prioritaskan xlsxwriter agar bisa insert checkbox
            try:
                with pd.ExcelWriter(out_xlsx, engine="xlsxwriter") as writer:  # type: ignore
                    df.to_excel(writer, index=False, sheet_name="Sheet1")
                    workbook  = writer.book
                    worksheet = writer.sheets["Sheet1"]
                    # Hitung dimensi data
                    last_row = len(df)
                    last_col = len(df.columns) - 1
                    # Terapkan autofilter (baris 0 adalah header)
                    worksheet.autofilter(0, 0, last_row, last_col)
                    # Freeze header
                    worksheet.freeze_panes(1, 0)
                    # Auto-resize kolom
                    for c, col_name in enumerate(df.columns):
                        series = df[col_name].astype(str)
                        max_len = max([len(col_name)] + [len(x) for x in series.tolist()])
                        width = max(10, min(60, max_len + 2))
                        worksheet.set_column(c, c, width)
            except Exception:
                # Fallback ke openpyxl jika xlsxwriter tidak ada
                try:
                    from openpyxl import load_workbook  # type: ignore
                    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:  # type: ignore
                        df.to_excel(writer, index=False, sheet_name="Sheet1")
                        ws = writer.book["Sheet1"]
                        # Set auto filter untuk seluruh area data
                        ws.auto_filter.ref = ws.dimensions
                        # Freeze header baris pertama
                        ws.freeze_panes = "A2"
                        # Auto-resize kolom berdasarkan panjang data
                        try:
                            from openpyxl.utils import get_column_letter  # type: ignore
                            for idx, col_name in enumerate(df.columns, start=1):
                                series = df[col_name].astype(str)
                                max_len = max([len(col_name)] + [len(x) for x in series.tolist()])
                                width = max(10, min(60, max_len + 2))
                                ws.column_dimensions[get_column_letter(idx)].width = width
                        except Exception:
                            pass
                except Exception:
                    # Jika kedua engine tidak tersedia, tulis tanpa fitur tambahan
                    df.to_excel(out_xlsx, index=False)
        except Exception as e:
            # Jika engine Excel (mis. openpyxl/xlsxwriter) belum terpasang, lanjutkan tanpa XLSX
            print(
                "Gagal menulis Excel (", e, ") -> Melewatkan XLSX. "
                "Install salah satu: 'pip install xlsxwriter' atau 'pip install openpyxl' untuk mengaktifkan ekspor Excel.")


def main():
    base = Path(__file__).parent
    input_csv = base / "jadwal-uts.csv"
    rooms_csv = base / "ruangan-kampus.csv"
    
    # Load rooms from CSV file
    global ALL_ROOMS
    ALL_ROOMS = load_rooms_from_csv(rooms_csv)
    print(f"Loaded {len(ALL_ROOMS)} rooms from {rooms_csv.name}")
    
    items = parse_csv(input_csv)
    assignments = build_schedule(items)
    out_csv = base / "jadwal-uts-output.csv"
    out_xlsx = base / "jadwal-uts-output.xlsx"
    write_outputs(assignments, out_csv, out_xlsx)
    print(f"Selesai. Output: {out_csv.name} dan {out_xlsx.name}")


if __name__ == "__main__":
    main()


