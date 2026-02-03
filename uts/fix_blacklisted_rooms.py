import csv
import random
from datetime import datetime, timedelta, time
from collections import defaultdict
from pathlib import Path

try:
    import pandas as pd
except Exception:
    pd = None

# ============================ Konfigurasi ============================
START_DATE = datetime(2025, 11, 3)
END_DATE = datetime(2025, 11, 7)
ALLOWED_SHIFT_STARTS = [time(7, 30), time(10, 0), time(13, 0), time(15, 30)]
SHIFT_DURATION_MIN = 120

# Blacklist ruangan berdasarkan hari
BLACKLIST_MON_WED_SUFFIXES = {"KTT 2.08", "KTT 2.07", "KTT 2.06", "KTT 2.05", "KTT 2.04"}
BLACKLIST_MON_FRI_SUFFIXES = {"KTT 2.09"}

ALL_ROOMS = []

def is_within_uts_week(date_dt: datetime) -> bool:
    return START_DATE.date() <= date_dt.date() <= END_DATE.date()

def is_room_blacklisted_on_date(room: str, date_dt: datetime) -> bool:
    """Return True jika ruangan diblacklist pada tanggal tersebut."""
    if not is_within_uts_week(date_dt):
        return False
    weekday = date_dt.weekday()
    for suf in BLACKLIST_MON_FRI_SUFFIXES:
        if room.endswith(suf) and weekday <= 4:
            return True
    for suf in BLACKLIST_MON_WED_SUFFIXES:
        if room.endswith(suf) and weekday <= 2:
            return True
    return False

def load_rooms_from_csv(rooms_csv_path: Path) -> list[str]:
    rooms = []
    try:
        with rooms_csv_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, delimiter=";")
            rows = list(reader)
        for row in rows[1:]:
            if len(row) > 0 and row[0].strip():
                room_name = row[0].strip()
                if room_name:
                    rooms.append(room_name)
    except Exception as e:
        print(f"Error loading rooms: {e}")
    return rooms

def parse_existing_datetime(hari: str, tanggal: str, shift: str) -> tuple[datetime, datetime] | None:
    if not tanggal or not shift:
        return None
    date_formats = ["%d-%b-%y", "%d/%m/%Y", "%d-%m-%Y"]
    date_dt = None
    for fmt in date_formats:
        try:
            date_dt = datetime.strptime(tanggal.strip(), fmt)
            break
        except Exception:
            continue
    if date_dt is None:
        return None
    parts = shift.split("-")
    if len(parts) != 2:
        return None
    start_s = parts[0].strip().replace(" ", "").replace("\t", "")
    end_s = parts[1].strip().replace(" ", "").replace("\t", "")
    try:
        h1, m1 = map(int, start_s.split("."))
        h2, m2 = map(int, end_s.split("."))
    except Exception:
        return None
    start_dt = datetime.combine(date_dt.date(), time(h1, m1))
    end_dt = datetime.combine(date_dt.date(), time(h2, m2))
    return start_dt, end_dt

def format_time_range(start_dt: datetime, end_dt: datetime) -> str:
    return f"{start_dt:%H.%M} - {end_dt:%H.%M}"

def weekday_name(dt: datetime) -> str:
    mapping = {
        0: "SENIN",
        1: "SELASA",
        2: "RABU",
        3: "KAMIS",
        4: "JUM'AT",
        5: "SABTU",
        6: "MINGGU",
    }
    return mapping[dt.weekday()]

def generate_daily_shifts(start_date: datetime) -> list[tuple[datetime, datetime]]:
    shifts = []
    for s in ALLOWED_SHIFT_STARTS:
        start_dt = datetime.combine(start_date.date(), s)
        end_dt = start_dt + timedelta(minutes=SHIFT_DURATION_MIN)
        shifts.append((start_dt, end_dt))
    return shifts

def iter_allowed_dates():
    cur = START_DATE
    while cur.date() <= END_DATE.date():
        if cur.weekday() < 5:
            yield cur
        cur += timedelta(days=1)

def pick_free_room(date_dt: datetime, date_key: str, shift_key: str, start_dt: datetime, end_dt: datetime, 
                   room_usage: dict, allow_aula: bool = False, bentuk_ujian: str = "", jumlah_mhs: int = 0) -> str | None:
    used_counts = room_usage.get(date_key, {}).get(shift_key, {})
    normal_candidates = []
    aula_candidates = []
    
    for r in ALL_ROOMS:
        if is_room_blacklisted_on_date(r, date_dt):
            continue
        
        used_count = used_counts.get(r, 0)
        
        if r.strip().upper() == "AULA":
            if not allow_aula:
                continue
            bentuk = (bentuk_ujian or "").strip().lower()
            if bentuk != "ujian tulis" or jumlah_mhs <= 0 or jumlah_mhs < 40:
                continue
            if used_count < 2:
                aula_candidates.append(r)
        else:
            if used_count == 0:
                normal_candidates.append(r)
    
    # Prioritize normal rooms first, then AULA if allowed
    if normal_candidates:
        return random.choice(normal_candidates)
    if aula_candidates:
        return random.choice(aula_candidates)
    return None

def main():
    base = Path(__file__).parent
    input_csv = base / "jadwal-uts-fix.csv"
    rooms_csv = base / "ruangan-kampus.csv"
    
    # Load rooms
    global ALL_ROOMS
    ALL_ROOMS = load_rooms_from_csv(rooms_csv)
    print(f"Loaded {len(ALL_ROOMS)} rooms")
    
    # Read all rows from CSV
    rows = []
    with input_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=";")
        header = next(reader)
        rows = list(reader)
    
    # Build column index - header first column is PROGRAM STUDI, not HARI
    col_idx = {name.strip().upper(): i for i, name in enumerate(header)}
    print(f"CSV columns: {list(col_idx.keys())}")
    
    def get(row, key, default=""):
        i = col_idx.get(key, None)
        if i is None or i >= len(row):
            return default
        return row[i].strip()
    
    # Build usage map from all non-blacklisted entries
    room_usage = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    class_usage = defaultdict(list)
    class_daily_count = defaultdict(lambda: defaultdict(int))
    
    blacklisted_indices = []
    
    # First pass: identify blacklisted entries and build usage map from others
    for idx, row in enumerate(rows):
        if not any(row):
            continue
        
        hari = get(row, "HARI")
        tanggal = get(row, "TANGGAL")
        shift = get(row, "SHIFT")
        ruangan = get(row, "RUANGAN")
        kelas = get(row, "KELAS")
        
        if not hari or not tanggal or not shift or not ruangan:
            # Empty entries - still add to usage if they have valid time/room
            if hari and tanggal and shift and ruangan:
                parsed = parse_existing_datetime(hari, tanggal, shift)
                if parsed:
                    start_dt, end_dt = parsed
                    date_key = start_dt.strftime("%Y-%m-%d")
                    shift_key = format_time_range(start_dt, end_dt)
                    if not is_room_blacklisted_on_date(ruangan, start_dt):
                        room_usage[date_key][shift_key][ruangan] += 1
                        if kelas:
                            class_usage[kelas].append((start_dt, end_dt))
                            class_daily_count[kelas][date_key] += 1
            continue
        
        parsed = parse_existing_datetime(hari, tanggal, shift)
        if parsed is None:
            continue
        
        start_dt, end_dt = parsed
        date_key = start_dt.strftime("%Y-%m-%d")
        shift_key = format_time_range(start_dt, end_dt)
        
        # Check if this room is blacklisted on this date
        if is_room_blacklisted_on_date(ruangan, start_dt):
            blacklisted_indices.append(idx)
            continue
        
        # Add to usage map (this entry is valid)
        room_usage[date_key][shift_key][ruangan] += 1
        if kelas:
            class_usage[kelas].append((start_dt, end_dt))
            class_daily_count[kelas][date_key] += 1
    
    print(f"Found {len(blacklisted_indices)} entries with blacklisted rooms")
    if not blacklisted_indices:
        print("No blacklisted rooms found. Generating Excel file anyway...")
        # Continue to generate Excel file even if no fixes needed
    
    # Second pass: regenerate blacklisted entries
    def is_class_conflict(kelas: str, start_dt: datetime, end_dt: datetime) -> bool:
        if not kelas:
            return False
        for s, e in class_usage.get(kelas, []):
            if not (end_dt <= s or start_dt >= e):
                return True
        date_key = start_dt.strftime("%Y-%m-%d")
        if class_daily_count[kelas][date_key] >= 2:
            return True
        return False
    
    fixed_count = 0
    if blacklisted_indices:
        for idx in blacklisted_indices:
            row = rows[idx]
            hari = get(row, "HARI")
            tanggal = get(row, "TANGGAL")
            shift = get(row, "SHIFT")
            kelas = get(row, "KELAS")
            bentuk_ujian = get(row, "BENTUK UJIAN")
            jumlah_mhs_str = get(row, "JUMLAH MAHASISWA")
            
            try:
                jumlah_mhs = int(jumlah_mhs_str) if jumlah_mhs_str else 0
            except:
                jumlah_mhs = 0
            
            parsed = parse_existing_datetime(hari, tanggal, shift)
            if parsed is None:
                print(f"Warning: Could not parse datetime for row {idx+1}")
                continue
            
            start_dt, end_dt = parsed
            date_key = start_dt.strftime("%Y-%m-%d")
            shift_key = format_time_range(start_dt, end_dt)
            
            # Try to find a new room for the same time slot first
            new_room = None
            allow_aula = (bentuk_ujian.strip().lower() == "ujian tulis" and jumlah_mhs >= 40)
            new_room = pick_free_room(start_dt, date_key, shift_key, start_dt, end_dt, 
                                     room_usage, allow_aula, bentuk_ujian, jumlah_mhs)
            
            found_new_slot = False
            
            # If found room at same time, just update the room
            if new_room:
                ruangan_col = col_idx.get("RUANGAN")
                if ruangan_col is not None:
                    while len(row) <= ruangan_col:
                        row.append("")
                    row[ruangan_col] = new_room
                    room_usage[date_key][shift_key][new_room] += 1
                    if kelas:
                        class_usage[kelas].append((start_dt, end_dt))
                        class_daily_count[kelas][date_key] += 1
                    fixed_count += 1
                    found_new_slot = True
                    print(f"Fixed row {idx+1}: Same time, new room {new_room}")
            
            # If no room found at same time, try different times on same date
            if not found_new_slot:
                for s_start, s_end in generate_daily_shifts(start_dt):
                    if is_class_conflict(kelas, s_start, s_end):
                        continue
                    date_key_new = s_start.strftime("%Y-%m-%d")
                    shift_key_new = format_time_range(s_start, s_end)
                    new_room = pick_free_room(s_start, date_key_new, shift_key_new, s_start, s_end,
                                             room_usage, allow_aula, bentuk_ujian, jumlah_mhs)
                    if new_room:
                        # Update row with new time and room
                        hari_col = col_idx.get("HARI")
                        tanggal_col = col_idx.get("TANGGAL")
                        shift_col = col_idx.get("SHIFT")
                        ruangan_col = col_idx.get("RUANGAN")
                        
                        # Ensure row has enough columns
                        max_col = max([c for c in [hari_col, tanggal_col, shift_col, ruangan_col] if c is not None])
                        while len(row) <= max_col:
                            row.append("")
                        
                        if hari_col is not None:
                            row[hari_col] = weekday_name(s_start)
                        if tanggal_col is not None:
                            row[tanggal_col] = s_start.strftime("%d-%b-%y")
                        if shift_col is not None:
                            row[shift_col] = shift_key_new
                        if ruangan_col is not None:
                            row[ruangan_col] = new_room
                        
                        # Update usage maps
                        room_usage[date_key_new][shift_key_new][new_room] += 1
                        if kelas:
                            class_usage[kelas].append((s_start, s_end))
                            class_daily_count[kelas][date_key_new] += 1
                        
                        fixed_count += 1
                        found_new_slot = True
                        print(f"Fixed row {idx+1}: New time {shift_key_new}, new room {new_room}")
                        break
            
            # If still no room, try different dates
            if not found_new_slot:
                for day_dt in iter_allowed_dates():
                    for s_start, s_end in generate_daily_shifts(day_dt):
                        if is_class_conflict(kelas, s_start, s_end):
                            continue
                        date_key_new = s_start.strftime("%Y-%m-%d")
                        shift_key_new = format_time_range(s_start, s_end)
                        new_room = pick_free_room(s_start, date_key_new, shift_key_new, s_start, s_end,
                                                 room_usage, allow_aula, bentuk_ujian, jumlah_mhs)
                        if new_room:
                            hari_col = col_idx.get("HARI")
                            tanggal_col = col_idx.get("TANGGAL")
                            shift_col = col_idx.get("SHIFT")
                            ruangan_col = col_idx.get("RUANGAN")
                            
                            max_col = max([c for c in [hari_col, tanggal_col, shift_col, ruangan_col] if c is not None])
                            while len(row) <= max_col:
                                row.append("")
                            
                            if hari_col is not None:
                                row[hari_col] = weekday_name(s_start)
                            if tanggal_col is not None:
                                row[tanggal_col] = s_start.strftime("%d-%b-%y")
                            if shift_col is not None:
                                row[shift_col] = shift_key_new
                            if ruangan_col is not None:
                                row[ruangan_col] = new_room
                            
                            room_usage[date_key_new][shift_key_new][new_room] += 1
                            if kelas:
                                class_usage[kelas].append((s_start, s_end))
                                class_daily_count[kelas][date_key_new] += 1
                            
                            fixed_count += 1
                            found_new_slot = True
                            print(f"Fixed row {idx+1}: New date {date_key_new}, time {shift_key_new}, room {new_room}")
                            break
                    if found_new_slot:
                        break
            
            if not found_new_slot:
                print(f"Warning: Could not find replacement for row {idx+1} (kelas: {kelas})")
    
    if blacklisted_indices:
        print(f"\nFixed {fixed_count} out of {len(blacklisted_indices)} entries")
    else:
        print(f"\nNo entries to fix.")
    
    # Write output CSV
    output_csv = base / "jadwal-uts-fix.csv"
    with output_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(header)
        for row in rows:
            writer.writerow(row)
    
    print(f"Output CSV written to {output_csv.name}")
    
    # Write output Excel dengan filter dan auto-adjust
    output_xlsx = base / "jadwal-uts-fix.xlsx"
    if pd is not None:
        try:
            # Konversi rows ke DataFrame
            data_rows = []
            for row in rows:
                # Pastikan row memiliki jumlah kolom yang sama dengan header
                padded_row = row + [""] * (len(header) - len(row))
                padded_row = padded_row[:len(header)]
                data_rows.append(padded_row)
            
            df = pd.DataFrame(data_rows, columns=header)
            
            # Prioritaskan xlsxwriter
            try:
                with pd.ExcelWriter(output_xlsx, engine="xlsxwriter") as writer:  # type: ignore
                    df.to_excel(writer, index=False, sheet_name="Sheet1")
                    workbook = writer.book
                    worksheet = writer.sheets["Sheet1"]
                    
                    # Hitung dimensi data
                    last_row = len(df)
                    last_col = len(df.columns) - 1
                    
                    # Terapkan autofilter pada header
                    worksheet.autofilter(0, 0, last_row, last_col)
                    
                    # Freeze header baris pertama
                    worksheet.freeze_panes(1, 0)
                    
                    # Format header: bold
                    header_format = workbook.add_format({"bold": True, "bg_color": "#D3D3D3"})
                    for col_num, col_name in enumerate(df.columns):
                        worksheet.write(0, col_num, col_name, header_format)
                    
                    # Auto-resize kolom dan set wrap text
                    wrap_format = workbook.add_format({"text_wrap": True, "valign": "top"})
                    col_widths = {}
                    for c, col_name in enumerate(df.columns):
                        series = df[col_name].astype(str)
                        # Hitung max length termasuk header
                        max_len = max([len(str(col_name))] + [len(str(x)) for x in series.tolist()])
                        width = max(10, min(60, max_len + 2))
                        col_widths[c] = width
                        worksheet.set_column(c, c, width, wrap_format)
                    
                    # Auto-adjust row height berdasarkan konten
                    for row_idx in range(1, last_row + 1):  # Mulai dari 1 (skip header)
                        max_lines = 1
                        for col_idx, col_name in enumerate(df.columns):
                            cell_value = str(df.iloc[row_idx - 1, col_idx])
                            if cell_value:
                                # Hitung jumlah baris berdasarkan panjang teks dan lebar kolom
                                col_width = col_widths.get(col_idx, 10)
                                text_length = len(cell_value)
                                # Estimasi: sekitar 8 karakter per unit lebar kolom
                                chars_per_line = max(8, int(col_width * 8))
                                lines = max(1, (text_length + chars_per_line - 1) // chars_per_line)
                                max_lines = max(max_lines, lines)
                        # Set tinggi baris (15 point per baris, minimum 15)
                        row_height = max(15, min(60, max_lines * 15))
                        worksheet.set_row(row_idx, row_height)
                    
                print(f"Excel file created successfully with xlsxwriter")
                    
            except Exception as e_xlsxwriter:
                # Fallback ke openpyxl jika xlsxwriter tidak ada
                print(f"xlsxwriter failed: {e_xlsxwriter}, trying openpyxl...")
                try:
                    from openpyxl import load_workbook  # type: ignore
                    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
                    from openpyxl.utils import get_column_letter  # type: ignore
                    
                    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:  # type: ignore
                        df.to_excel(writer, index=False, sheet_name="Sheet1")
                        ws = writer.book["Sheet1"]
                        
                        # Set auto filter untuk seluruh area data
                        ws.auto_filter.ref = ws.dimensions
                        
                        # Freeze header baris pertama
                        ws.freeze_panes = "A2"
                        
                        # Format header: bold dan background abu-abu
                        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        header_font = Font(bold=True)
                        for cell in ws[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Auto-resize kolom dan set wrap text
                        wrap_alignment = Alignment(wrap_text=True, vertical="top")
                        for idx, col_name in enumerate(df.columns, start=1):
                            series = df[col_name].astype(str)
                            max_len = max([len(str(col_name))] + [len(str(x)) for x in series.tolist()])
                            width = max(10, min(60, max_len + 2))
                            col_letter = get_column_letter(idx)
                            ws.column_dimensions[col_letter].width = width
                            
                            # Set wrap text untuk semua cell di kolom ini
                            for row_idx in range(2, len(df) + 2):  # Mulai dari baris 2 (skip header)
                                cell = ws[f"{col_letter}{row_idx}"]
                                cell.alignment = wrap_alignment
                        
                        # Auto-adjust row height berdasarkan konten
                        for row_idx in range(2, len(df) + 2):  # Mulai dari baris 2 (skip header)
                            max_lines = 1
                            for col_idx, col_name in enumerate(df.columns, start=1):
                                cell_value = str(df.iloc[row_idx - 2, col_idx - 1])
                                if cell_value:
                                    col_letter = get_column_letter(col_idx)
                                    col_width = ws.column_dimensions[col_letter].width
                                    text_length = len(cell_value)
                                    # Estimasi: sekitar 7 karakter per unit lebar kolom
                                    chars_per_line = max(7, int(col_width * 7))
                                    lines = max(1, (text_length + chars_per_line - 1) // chars_per_line)
                                    max_lines = max(max_lines, lines)
                            # Set tinggi baris dalam points (15 points per baris, minimum 15)
                            row_height = max(15, min(60, max_lines * 15))
                            ws.row_dimensions[row_idx].height = row_height
                        
                    print(f"Excel file created successfully with openpyxl")
                        
                except Exception as e_openpyxl:
                    # Jika kedua engine tidak tersedia, tulis tanpa fitur tambahan
                    print(f"openpyxl failed: {e_openpyxl}, trying basic export...")
                    try:
                        df.to_excel(output_xlsx, index=False)
                        print(f"Warning: Excel written without formatting")
                    except Exception as e_basic:
                        print(f"Basic Excel export also failed: {e_basic}")
        except Exception as e:
            import traceback
            print(f"Gagal menulis Excel: {e}")
            print(traceback.format_exc())
            print("Install salah satu: 'pip install xlsxwriter' atau 'pip install openpyxl' untuk mengaktifkan ekspor Excel.")
    else:
        print("Pandas tidak tersedia. Hanya file CSV yang dibuat.")
    
    if output_xlsx.exists():
        print(f"Output Excel written to {output_xlsx.name}")

if __name__ == "__main__":
    main()

